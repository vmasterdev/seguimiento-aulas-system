#!/usr/bin/env python3

import json
import os
import sys

from openpyxl import load_workbook


def main() -> int:
    if len(sys.argv) != 2:
        print("Uso: fill_auditor_template.py <payload.json>", file=sys.stderr)
        return 1

    payload_path = sys.argv[1]
    with open(payload_path, "r", encoding="utf-8") as handle:
        payload = json.load(handle)

    template_path = payload["templatePath"]
    output_path = payload["outputPath"]
    identity = payload["identity"]
    rows = payload["rows"]

    workbook = load_workbook(template_path)
    worksheet = workbook[workbook.sheetnames[0]]

    start_row = 5
    for index, row in enumerate(rows, start=start_row):
        worksheet[f"B{index}"] = identity["firstName"]
        worksheet[f"C{index}"] = identity["lastName"]
        worksheet[f"D{index}"] = identity["institutionalEmail"]
        worksheet[f"E{index}"] = row["courseLabel"]
        worksheet[f"F{index}"] = identity["roleName"]

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    workbook.save(output_path)
    print(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
