#!/usr/bin/env python3
import csv
import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

INPUT_DIR = Path("1.2 CATEGORIZACION/2026 S1")
OUTPUT_XLSX = INPUT_DIR / "RESULTADO_CATEGORIZACION_AULAS.xlsx"


@dataclass
class NRCRecord:
    nrc: str
    periodos: set = field(default_factory=set)
    titulos: set = field(default_factory=set)
    metodos: set = field(default_factory=set)
    archivos: set = field(default_factory=set)
    inscritos_max: int = 0


def clean_text(value: str) -> str:
    return (value or "").strip()


def parse_int(value: str) -> int:
    value = clean_text(value)
    if not value:
        return 0
    try:
        return int(value)
    except ValueError:
        return 0


def normalize_nrc(raw: str) -> str:
    digits = re.sub(r"\D", "", raw or "")
    if len(digits) >= 5:
        return digits[-5:]
    return digits.zfill(5) if digits else ""


def classify(record: NRCRecord):
    titles_blob = " | ".join(sorted(record.titulos))
    files_upper = [f.upper() for f in record.archivos]
    methods_upper = {m.upper() for m in record.metodos}

    # Heuristica 1: cursos de opcion/proyecto de grado suelen ser aulas vacias.
    if re.search(r"\b(opci[oó]n\s+de\s+grado|proyecto\s+de\s+grado)\b", titles_blob, re.IGNORECASE):
        return "vacia", "media", "titulo_grado"

    # Heuristica 2: cursos con metodo DIST o archivos de modalidad distancia.
    if "DIST" in methods_upper or any("DISTANCIA" in f for f in files_upper):
        return "distancia 4.0", "media-baja", "metodo_dist_o_archivo_distancia"

    # Heuristica 3: posgrado presencial suele usar formato innovame.
    if any("POSGRADO PRESENCIAL" in f for f in files_upper):
        return "innovame", "baja", "archivo_posgrado_presencial"

    # Heuristica 4: resto se clasifica como criba.
    return "criba", "baja", "regla_por_descarte"


def load_records(input_dir: Path):
    records = defaultdict(lambda: NRCRecord(nrc=""))
    csv_files = sorted(p for p in input_dir.glob("*.csv") if not p.name.startswith("~$"))

    for csv_file in csv_files:
        with csv_file.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh, delimiter=";")
            for row in reader:
                nrc = normalize_nrc(row.get("NRC", ""))
                if not nrc:
                    continue
                rec = records[nrc]
                rec.nrc = nrc
                rec.periodos.add(clean_text(row.get("PERIODO", "")))
                title = clean_text(row.get("TITULO", ""))
                if title:
                    rec.titulos.add(title)
                method = clean_text(row.get("METODO_EDUCATIVO", ""))
                if method:
                    rec.metodos.add(method)
                rec.archivos.add(csv_file.name)
                rec.inscritos_max = max(rec.inscritos_max, parse_int(row.get("INSCRITOS", "0")))

    return records, csv_files


def build_workbook(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorizacion"

    headers = [
        "NRC",
        "CATEGORIA_AULA",
        "CONFIANZA",
        "REGLA_USADA",
        "VALIDACION_MANUAL",
        "PERIODOS",
        "METODO_EDUCATIVO",
        "INSCRITOS_MAX",
        "TITULO_REFERENCIA",
        "ARCHIVOS_ORIGEN",
    ]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = max(18, len(title) + 2)

    for nrc in sorted(records.keys(), key=lambda x: int(x)):
        rec = records[nrc]
        categoria, confianza, regla = classify(rec)

        periodos = ", ".join(sorted(p for p in rec.periodos if p))
        metodos = ", ".join(sorted(rec.metodos))
        titulo_ref = sorted(rec.titulos)[0] if rec.titulos else ""
        archivos = " | ".join(sorted(rec.archivos))

        ws.append(
            [
                rec.nrc,
                categoria,
                confianza,
                regla,
                "SI",
                periodos,
                metodos,
                rec.inscritos_max,
                titulo_ref,
                archivos,
            ]
        )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Congelar encabezado
    ws.freeze_panes = "A2"

    return wb


def main():
    if not INPUT_DIR.exists():
        raise SystemExit(f"No existe el directorio de entrada: {INPUT_DIR}")

    records, csv_files = load_records(INPUT_DIR)
    if not records:
        raise SystemExit("No se encontraron NRC para procesar.")

    wb = build_workbook(records)
    wb.save(OUTPUT_XLSX)

    print(f"CSV procesados: {len(csv_files)}")
    print(f"NRC unicos: {len(records)}")
    print(f"Excel generado: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
