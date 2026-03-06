import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import csv
from difflib import SequenceMatcher
import re
import shutil
import time
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Set, Tuple
from urllib.parse import parse_qs, urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options as ChromeOptions
    from selenium.webdriver.edge.options import Options as EdgeOptions
    from selenium.webdriver.common.by import By
    from selenium.webdriver.remote.webdriver import WebDriver
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait
    SELENIUM_IMPORT_ERROR = None
except Exception as e:
    webdriver = None
    ChromeOptions = None
    EdgeOptions = None
    By = None
    WebDriver = object
    EC = None
    WebDriverWait = None
    SELENIUM_IMPORT_ERROR = e


# =========================
# CONFIGURACION
# =========================
PROFILE_DIR_CHROME = str(Path.cwd() / "chrome_profile_aulas_uniminuto")
PROFILE_DIR_EDGE = str(Path.cwd() / "edge_profile_aulas_uniminuto")
DEFAULT_INPUT_DIR = Path("1.2 CATEGORIZACION/2026 S1")
DEFAULT_OUTPUT_XLSX = DEFAULT_INPUT_DIR / "RESULTADO_TIPOS_AULA_DESDE_MOODLE.xlsx"
VACIA_MAX_MOD_LINKS = 2
LOGIN_WAIT_SECONDS = 600
MAX_REINTENTOS_NRC = 3
MAX_WORKERS = 8
TITLE_SIMILAR_THRESHOLD = 80.0
TITLE_EQUAL_THRESHOLD = 99.0
D40_NAME_THRESHOLD = 86.0

PROFILE_COPY_IGNORE = (
    "Crashpad",
    "Singleton*",
    "LOCK",
    "lockfile",
    "*.log",
    "*.ldb",
    "*Cache*",
    "GPUCache",
    "Code Cache",
    "ShaderCache",
    "GrShaderCache",
    "BrowserMetrics",
)

MODALIDADES = {
    "presencial": {
        "nombre": "PRESENCIAL",
        "base_url": "https://presencial.aulasuniminuto.edu.co",
        "prefixes": ("60",),
    },
    "posgrados": {
        "nombre": "POSGRADOS",
        "base_url": "https://posgrados.aulasuniminuto.edu.co",
        "prefixes": ("61", "71"),
    },
    "moocs": {
        "nombre": "MOOCs",
        "base_url": "https://moocs.aulasuniminuto.edu.co",
        "prefixes": ("62", "86"),
    },
    "distancia": {
        "nombre": "DISTANCIA",
        "base_url": "https://distancia.aulasuniminuto.edu.co",
        "prefixes": ("65",),
    },
}

TODAS_MODALIDADES = list(MODALIDADES.keys())

# Prioridad por codigo de periodo (anio + indicador) definido por el usuario.
PERIODO_MODALIDADES = {
    "202610": ["presencial"],  # pregrado presencial
    "202660": ["presencial"],  # pregrado presencial
    "202615": ["distancia"],  # pregrado distancia
    "202665": ["distancia"],  # pregrado distancia
    "202611": ["posgrados"],  # posgrados
    "202661": ["posgrados"],  # posgrados
    "202621": ["posgrados"],  # posgrados
    "202671": ["posgrados"],  # posgrados
    "202641": ["posgrados"],  # cuatrimestral
    "202580": ["moocs"],  # intersemestral
}

NO_MATRICULADO_PATTERNS = [
    "no esta matriculado en este curso",
    "no estas matriculado en este curso",
    "no se puede auto matricular en este curso",
    "no se puede automatricular en este curso",
    "you are not enrolled in this course",
    "not enrolled in this course",
    "no tiene permisos para ver este curso",
]

PARTICIPANTES_CERO_PATTERNS = [
    "0 participantes encontrados",
    "0 participante encontrado",
    "0 participants found",
    "0 participant found",
]

LOGIN_URL_HINTS = (
    "/login/",
    "/auth/oidc",
    "/oauth2/",
    "login.microsoftonline.com",
    "microsoftonline.com",
)

DEFAULT_CATALOGO_D40_XLSX = Path(__file__).resolve().parent / "CURSOS DISTANCIA 4.0.xlsx"
DEFAULT_CATALOGO_D40_HOJA = "DISTANCIA 4.0"
DEFAULT_CATALOGO_D40_COL = 3  # columna C
ARCHIVOS_SOLO_NRC_5_HINTS = (
    "PREGRADO DISTANCIA",
    "PREGRADO PRESENCIAL",
    "POSGRADO PRESENCIAL",
    "POSGRADO DISTANCIA",
)

OUTPUT_HEADERS = [
    "NRC",
    "TITULO_FUENTE",
    "TITULOS_FUENTE",
    "TIPO_AULA",
    "TOTAL_PARTICIPANTES",
    "PARTICIPANTES_DETECTADOS",
    "ES_VACIA",
    "ES_VACIA_POR_PARTICIPANTES",
    "ES_VACIA_AMBOS_CRITERIOS",
    "TEXTO_PARTICIPANTES",
    "ES_INNOVAME",
    "ES_DISTANCIA_4_VISUAL",
    "CONFIANZA",
    "REGLA_DETECCION",
    "MOD_LINKS",
    "MODALIDAD_DONDE_SE_ENCONTRO",
    "COURSE_ID",
    "NOMBRE_CURSO_MOODLE",
    "COINCIDENCIA_TITULO",
    "SCORE_COINCIDENCIA_TITULO",
    "TITULO_CATALOGO_D40_MATCH",
    "SCORE_CATALOGO_D40",
    "ES_DISTANCIA_4_POR_NOMBRE",
    "ES_DISTANCIA_4_AMBOS",
    "CONSULTA_USADA",
    "PERIODOS",
    "METODOS",
    "ARCHIVOS_ORIGEN",
    "ESTADO",
    "ERROR",
]


@dataclass
class NRCRecord:
    nrc: str
    periodos: Set[str] = field(default_factory=set)
    titulos: Set[str] = field(default_factory=set)
    metodos: Set[str] = field(default_factory=set)
    archivos: Set[str] = field(default_factory=set)
    modalidades_preferidas: List[str] = field(default_factory=list)


def requiere_solo_nrc_5_por_archivo(rec: NRCRecord) -> bool:
    for archivo in rec.archivos:
        nombre = (archivo or "").upper()
        if any(hint in nombre for hint in ARCHIVOS_SOLO_NRC_5_HINTS):
            return True
    return False


def unique_keep_order(values: Sequence[str]) -> List[str]:
    out: List[str] = []
    seen = set()
    for value in values:
        if not value:
            continue
        if value not in seen:
            out.append(value)
            seen.add(value)
    return out


def normalize_nrc(raw_nrc: str) -> str:
    digits = re.sub(r"\D", "", (raw_nrc or "").strip())
    if not digits:
        return ""
    if len(digits) > 5:
        digits = digits[-5:]
    return digits.zfill(5)


def normalize_periodo(raw_periodo: str) -> str:
    digits = re.sub(r"\D", "", (raw_periodo or "").strip())
    if len(digits) >= 6:
        return digits[:6]
    return digits


def strip_leading_zeros(nrc: str) -> str:
    value = nrc.lstrip("0")
    return value if value else "0"


def _sufijo_periodo(periodo: str) -> str:
    periodo_norm = normalize_periodo(periodo)
    if len(periodo_norm) >= 2:
        return periodo_norm[-2:]
    return ""


def construir_nrc_queries(rec: NRCRecord, nrc: str, usar_solo_nrc_5: bool) -> List[str]:
    nrc_5 = normalize_nrc(nrc)
    if not nrc_5:
        return []

    nrc_sin_ceros = strip_leading_zeros(nrc_5)
    sufijos_periodo = unique_keep_order(_sufijo_periodo(p) for p in rec.periodos if p)
    queries: List[str] = []

    # Prioriza busquedas del tipo "80-380" / "80-00380" para reducir ambiguedades.
    for sufijo in sufijos_periodo:
        if not sufijo:
            continue
        if usar_solo_nrc_5:
            queries.append(f"{sufijo}-{nrc_5}")
        else:
            queries.append(f"{sufijo}-{nrc_sin_ceros}")
            queries.append(f"{sufijo}-{nrc_5}")

    if usar_solo_nrc_5:
        queries.append(nrc_5)
    else:
        queries.append(nrc_sin_ceros)
        queries.append(nrc_5)
    return unique_keep_order(queries)


def normalize_title(value: str) -> str:
    text = (value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def similarity_score(a: str, b: str) -> float:
    a_norm = normalize_title(a)
    b_norm = normalize_title(b)
    if not a_norm or not b_norm:
        return 0.0
    if a_norm == b_norm:
        return 100.0

    ratio_direct = SequenceMatcher(None, a_norm, b_norm).ratio()
    a_tokens = " ".join(sorted(a_norm.split()))
    b_tokens = " ".join(sorted(b_norm.split()))
    ratio_tokens = SequenceMatcher(None, a_tokens, b_tokens).ratio()
    ratio_contiene = 0.0
    if (a_norm in b_norm or b_norm in a_norm) and min(len(a_norm), len(b_norm)) >= 8:
        ratio_contiene = 0.97
    return round(max(ratio_direct, ratio_tokens, ratio_contiene) * 100.0, 1)


def format_score(value: float) -> str:
    return f"{value:.1f}"


def bool_to_si_no(value: bool) -> str:
    return "SI" if value else "NO"


def elegir_titulo_principal(titulos: Set[str]) -> str:
    vals = [(t or "").strip() for t in titulos if (t or "").strip()]
    if not vals:
        return ""
    vals = sorted(vals, key=lambda x: (-len(normalize_title(x)), x.lower()))
    return vals[0]


def comparar_titulos_fuente_vs_moodle(
    titulos_fuente: Sequence[str], titulo_moodle: str
) -> Tuple[str, float]:
    titulo_moodle = (titulo_moodle or "").strip()
    if not titulo_moodle:
        return "SIN_MOODLE", 0.0
    candidatos = [(t or "").strip() for t in titulos_fuente if (t or "").strip()]
    if not candidatos:
        return "SIN_TITULO_FUENTE", 0.0

    best = 0.0
    for src in candidatos:
        best = max(best, similarity_score(src, titulo_moodle))

    if best >= TITLE_EQUAL_THRESHOLD:
        return "IGUAL", best
    if best >= TITLE_SIMILAR_THRESHOLD:
        return "SIMILAR", best
    return "DIFERENTE", best


def cargar_catalogo_distancia40(
    xlsx_path: Path,
    sheet_name: str = DEFAULT_CATALOGO_D40_HOJA,
    col_index: int = DEFAULT_CATALOGO_D40_COL,
) -> List[str]:
    if not xlsx_path.exists():
        print(f"[WARN] No existe catalogo Distancia 4.0: {xlsx_path}")
        return []
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"[WARN] No se pudo abrir catalogo Distancia 4.0 ({xlsx_path}): {e}")
        return []

    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    out: List[str] = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_index, max_col=col_index, values_only=True):
        raw = row[0] if row else None
        value = "" if raw is None else str(raw).strip()
        if value:
            out.append(value)
    return unique_keep_order(out)


def mejor_match_catalogo_d40(candidatos: Sequence[str], catalogo: Sequence[str]) -> Tuple[str, float]:
    catalogo_vals = [(c or "").strip() for c in catalogo if (c or "").strip()]
    candidatos_vals = [(c or "").strip() for c in candidatos if (c or "").strip()]
    if not catalogo_vals or not candidatos_vals:
        return "", 0.0

    best_title = ""
    best_score = 0.0
    for cand in candidatos_vals:
        cand_norm = normalize_title(cand)
        if not cand_norm:
            continue
        for cat_title in catalogo_vals:
            score = similarity_score(cand, cat_title)
            if score > best_score:
                best_score = score
                best_title = cat_title
    return best_title, best_score


def build_enriched_row(
    rec: NRCRecord,
    nrc: str,
    tipo_aula: str,
    total_participantes: Optional[int],
    es_vacia_por_participantes: bool,
    texto_participantes: str,
    confianza: str,
    regla: str,
    mod_links: str,
    modalidad_encontrada: str,
    course_id: str,
    nombre_curso_moodle: str,
    consulta_usada: str,
    periodos: str,
    metodos: str,
    archivos: str,
    estado: str,
    error: str,
    catalogo_d40: Sequence[str],
) -> Dict[str, str]:
    titulo_moodle = (nombre_curso_moodle or "").strip()
    titulos_fuente = sorted(t for t in rec.titulos if (t or "").strip())
    titulo_fuente = elegir_titulo_principal(rec.titulos)
    titulos_fuente_join = " | ".join(titulos_fuente)

    comparacion_titulo, score_titulo = comparar_titulos_fuente_vs_moodle(
        titulos_fuente=titulos_fuente,
        titulo_moodle=titulo_moodle,
    )
    match_d40, score_d40 = mejor_match_catalogo_d40(
        candidatos=[titulo_moodle, titulo_fuente],
        catalogo=catalogo_d40,
    )

    tipo_norm = (tipo_aula or "").strip().lower()
    es_vacia = tipo_norm == "vacia"
    es_vacia_ambos = es_vacia and es_vacia_por_participantes
    es_innovame = tipo_norm == "innovame"
    es_d40_visual = tipo_norm == "distancia 4.0"
    es_d40_nombre = score_d40 >= D40_NAME_THRESHOLD
    es_d40_ambos = es_d40_visual and es_d40_nombre

    return {
        "NRC": nrc,
        "TITULO_FUENTE": titulo_fuente,
        "TITULOS_FUENTE": titulos_fuente_join,
        "TIPO_AULA": tipo_aula,
        "TOTAL_PARTICIPANTES": "" if total_participantes is None else str(total_participantes),
        "PARTICIPANTES_DETECTADOS": bool_to_si_no(total_participantes is not None),
        "ES_VACIA": bool_to_si_no(es_vacia),
        "ES_VACIA_POR_PARTICIPANTES": bool_to_si_no(es_vacia_por_participantes),
        "ES_VACIA_AMBOS_CRITERIOS": bool_to_si_no(es_vacia_ambos),
        "TEXTO_PARTICIPANTES": texto_participantes,
        "ES_INNOVAME": bool_to_si_no(es_innovame),
        "ES_DISTANCIA_4_VISUAL": bool_to_si_no(es_d40_visual),
        "CONFIANZA": confianza,
        "REGLA_DETECCION": regla,
        "MOD_LINKS": mod_links,
        "MODALIDAD_DONDE_SE_ENCONTRO": modalidad_encontrada,
        "COURSE_ID": course_id,
        "NOMBRE_CURSO_MOODLE": titulo_moodle,
        "COINCIDENCIA_TITULO": comparacion_titulo,
        "SCORE_COINCIDENCIA_TITULO": format_score(score_titulo),
        "TITULO_CATALOGO_D40_MATCH": match_d40,
        "SCORE_CATALOGO_D40": format_score(score_d40),
        "ES_DISTANCIA_4_POR_NOMBRE": bool_to_si_no(es_d40_nombre),
        "ES_DISTANCIA_4_AMBOS": bool_to_si_no(es_d40_ambos),
        "CONSULTA_USADA": consulta_usada,
        "PERIODOS": periodos,
        "METODOS": metodos,
        "ARCHIVOS_ORIGEN": archivos,
        "ESTADO": estado,
        "ERROR": error,
    }


def profile_dir_base(browser: str) -> Path:
    browser_key = (browser or "edge").strip().lower()
    if browser_key == "edge":
        return Path(PROFILE_DIR_EDGE)
    if browser_key == "chrome":
        return Path(PROFILE_DIR_CHROME)
    raise ValueError("Navegador no soportado. Usa --browser edge o --browser chrome")


def profile_dir_worker(browser: str, worker_id: int) -> Path:
    base = profile_dir_base(browser)
    return base.parent / f"{base.name}_w{worker_id}"


def preparar_perfiles_workers(browser: str, workers: int) -> Dict[int, str]:
    workers = max(1, workers)
    if workers == 1:
        return {1: str(profile_dir_base(browser))}

    base = profile_dir_base(browser)
    if not base.exists():
        print(
            f"[WARN] No existe perfil base {base}. "
            "Se crearan perfiles vacios por worker y podria requerirse login."
        )

    perfiles: Dict[int, str] = {}
    for worker_id in range(1, workers + 1):
        worker_profile = profile_dir_worker(browser, worker_id)
        try:
            if worker_profile.exists():
                shutil.rmtree(worker_profile, ignore_errors=True)
            if base.exists():
                print(f"[INFO] Preparando perfil worker {worker_id}: {worker_profile}")
                shutil.copytree(
                    base,
                    worker_profile,
                    ignore=shutil.ignore_patterns(*PROFILE_COPY_IGNORE),
                    dirs_exist_ok=True,
                )
            else:
                worker_profile.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            print(
                f"[WARN] No se pudo preparar perfil para worker {worker_id} "
                f"({worker_profile}): {e}"
            )
            worker_profile.mkdir(parents=True, exist_ok=True)
        perfiles[worker_id] = str(worker_profile)
    return perfiles


def url_indica_login(url: str) -> bool:
    url_norm = (url or "").strip().lower()
    if not url_norm:
        return True
    return any(hint in url_norm for hint in LOGIN_URL_HINTS)


def sesion_activa_en_modalidad(current_url: str, base_url: str) -> bool:
    url_norm = (current_url or "").strip().lower()
    base_norm = (base_url or "").strip().lower()
    if not url_norm or not base_norm:
        return False
    if base_norm not in url_norm:
        return False
    return not url_indica_login(url_norm)


def obtener_id_desde_href(href: str) -> Optional[str]:
    parsed = urlparse(href)
    qs = parse_qs(parsed.query)
    return qs.get("id", [None])[0]


def inferir_modalidades(nrc: str, nombre_archivo: str, metodo: str, periodo: str) -> List[str]:
    prefs: List[str] = []

    archivo_upper = nombre_archivo.upper()
    metodo_upper = (metodo or "").upper()

    # 1) Prioridad por nombre de archivo (regla mas explicita).
    if "PREGRADO PRESENCIAL" in archivo_upper:
        prefs.extend(["presencial"])
    elif "POSGRADO PRESENCIAL" in archivo_upper:
        prefs.extend(["posgrados"])
    elif "PREGRADO DISTANCIA" in archivo_upper:
        prefs.extend(["distancia"])
    elif "POSGRADO DISTANCIA" in archivo_upper:
        prefs.extend(["posgrados"])
    elif "TECNICO LABORAL" in archivo_upper:
        prefs.extend(["moocs", "distancia", "presencial"])
    elif "CUATRIMESTRAL" in archivo_upper:
        prefs.extend(["posgrados"])
    elif "INTERSEMESTRAL" in archivo_upper:
        prefs.extend(["moocs"])
    elif "OPCION DE GRADO" in archivo_upper:
        prefs.extend(["distancia", "presencial", "posgrados", "moocs"])

    # 2) Prioridad por PERIODO (anio + indicador) segun reglas de negocio.
    periodo_norm = normalize_periodo(periodo)
    if periodo_norm in PERIODO_MODALIDADES:
        prefs.extend(PERIODO_MODALIDADES[periodo_norm])

    # 3) Refuerzo por prefijo NRC.
    for clave, cfg in MODALIDADES.items():
        if any(nrc.startswith(pref) for pref in cfg["prefixes"]):
            prefs.append(clave)
            break

    # Refuerzo por metodo educativo
    if metodo_upper == "DIST":
        prefs.insert(0, "distancia")

    return unique_keep_order(prefs)


def parse_modalidades_cli(raw_value: str) -> List[str]:
    raw = (raw_value or "").strip()
    if not raw:
        return []

    alias_to_key: Dict[str, str] = {}
    for key, cfg in MODALIDADES.items():
        alias_to_key[key.lower()] = key
        alias_to_key[str(cfg.get("nombre", "")).strip().lower()] = key

    result: List[str] = []
    invalid: List[str] = []
    tokens = re.split(r"[,;|/\s]+", raw)
    for token in tokens:
        value = (token or "").strip().lower()
        if not value:
            continue
        key = alias_to_key.get(value)
        if not key:
            invalid.append(token)
            continue
        result.append(key)

    if invalid:
        valid_values = ", ".join(sorted(set(alias_to_key.keys())))
        raise ValueError(
            "Modalidades no validas: "
            + ", ".join(invalid)
            + f". Valores permitidos: {valid_values}"
        )

    return unique_keep_order(result)


def detectar_delimitador_csv(csv_file: Path) -> str:
    try:
        sample = csv_file.read_text(encoding="utf-8-sig", errors="ignore")[:8192]
    except Exception:
        return ";"
    if not sample.strip():
        return ";"
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        if dialect.delimiter in (";", ","):
            return dialect.delimiter
    except Exception:
        pass
    return "," if sample.count(",") > sample.count(";") else ";"


def _norm_col(name: str) -> str:
    text = (name or "").strip().replace('"', "")
    text = re.sub(r"\s+", "", text)
    return text.upper()


def valor_columna(row: Dict[str, str], *candidatos: str) -> str:
    for c in candidatos:
        if c in row and row.get(c) is not None:
            return str(row.get(c) or "").strip()
    norm_map = {_norm_col(k): ("" if v is None else str(v).strip()) for k, v in row.items()}
    for c in candidatos:
        key = _norm_col(c)
        if key in norm_map:
            return norm_map[key]
    return ""


def iter_rows_csv(csv_file: Path):
    delimiter = detectar_delimitador_csv(csv_file)
    with csv_file.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh, delimiter=delimiter)
        headers = reader.fieldnames or []
        # Fallback adicional por si el delimitador detectado no funciono.
        if len(headers) == 1 and "," in (headers[0] or "") and delimiter != ",":
            fh.seek(0)
            reader = csv.DictReader(fh, delimiter=",")
        elif len(headers) == 1 and ";" in (headers[0] or "") and delimiter != ";":
            fh.seek(0)
            reader = csv.DictReader(fh, delimiter=";")
        for row in reader:
            if row:
                yield row


def leer_headers_input_csv(csv_file: Path) -> List[str]:
    delimiter = detectar_delimitador_csv(csv_file)
    with csv_file.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh, delimiter=delimiter)
        headers = reader.fieldnames or []
        # Fallback adicional por si el delimitador detectado no funciono.
        if len(headers) == 1 and "," in (headers[0] or "") and delimiter != ",":
            fh.seek(0)
            headers = (csv.DictReader(fh, delimiter=",").fieldnames or [])
        elif len(headers) == 1 and ";" in (headers[0] or "") and delimiter != ";":
            fh.seek(0)
            headers = (csv.DictReader(fh, delimiter=";").fieldnames or [])
    return [h for h in headers if h is not None]


def es_csv_salida_categorizacion(csv_file: Path, headers: Optional[List[str]] = None) -> bool:
    nombre = (csv_file.name or "").upper()
    if nombre.startswith("RESULTADO_TIPOS_AULA_DESDE_MOODLE"):
        return True

    raw_headers = headers if headers is not None else leer_headers_input_csv(csv_file)
    norm_headers = {_norm_col(h) for h in raw_headers if h}
    signature = {"NRC", "TIPOAULA", "ESTADO", "PERIODOS"}
    return signature.issubset(norm_headers)


def cargar_nrcs(input_dir: Path) -> Dict[str, NRCRecord]:
    records: Dict[str, NRCRecord] = {}
    csv_files = sorted(p for p in input_dir.glob("*.csv") if not p.name.startswith("~$"))
    if not csv_files:
        raise RuntimeError(f"No se encontraron CSV en {input_dir}")

    for csv_file in csv_files:
        headers = leer_headers_input_csv(csv_file)
        norm_headers = {_norm_col(h) for h in headers if h}
        if "NRC" not in norm_headers:
            print(f"[INFO] Omitiendo CSV sin columna NRC: {csv_file.name}")
            continue
        if es_csv_salida_categorizacion(csv_file, headers):
            print(f"[INFO] Omitiendo CSV de salida/resultado: {csv_file.name}")
            continue

        for row in iter_rows_csv(csv_file):
            nrc = normalize_nrc(valor_columna(row, "NRC"))
            if not nrc:
                continue
            if nrc not in records:
                records[nrc] = NRCRecord(nrc=nrc)
            rec = records[nrc]
            periodo = valor_columna(row, "PERIODO")
            titulo = valor_columna(row, "TITULO")
            metodo = valor_columna(row, "METODO_EDUCATIVO")
            rec.periodos.add(periodo)
            if titulo:
                rec.titulos.add(titulo)
            if metodo:
                rec.metodos.add(metodo)
            rec.archivos.add(csv_file.name)
            rec.modalidades_preferidas = unique_keep_order(
                rec.modalidades_preferidas
                + inferir_modalidades(nrc, csv_file.name, metodo, periodo)
            )

    return records


def create_driver(
    browser: str = "edge",
    clean_profile: bool = False,
    headless: bool = False,
    profile_dir_override: Optional[str] = None,
) -> WebDriver:
    if SELENIUM_IMPORT_ERROR is not None:
        raise RuntimeError(
            "No se pudo importar selenium. Instala dependencias con: "
            "pip install selenium openpyxl"
        ) from SELENIUM_IMPORT_ERROR

    browser_key = (browser or "edge").strip().lower()
    if browser_key == "edge":
        profile_dir = profile_dir_override or PROFILE_DIR_EDGE
        browser_name = "Edge"
    elif browser_key == "chrome":
        profile_dir = profile_dir_override or PROFILE_DIR_CHROME
        browser_name = "Chrome"
    else:
        raise ValueError("Navegador no soportado. Usa --browser edge o --browser chrome")

    if clean_profile:
        try:
            if Path(profile_dir).exists():
                print(f"[INFO] Borrando perfil de {browser_name} corrupto...")
                shutil.rmtree(profile_dir, ignore_errors=True)
        except Exception as e:
            print(f"[WARN] No se pudo borrar el perfil: {e}")

    if browser_key == "edge":
        options = EdgeOptions()
        options.add_argument(f"--user-data-dir={profile_dir}")
        options.add_argument("--no-first-run")
        options.add_argument("--no-default-browser-check")
        if headless:
            options.add_argument("--headless=new")
            options.add_argument("--disable-gpu")
            options.add_argument("--window-size=1920,1080")
            options.add_argument("--disable-dev-shm-usage")
        driver = webdriver.Edge(options=options)
    else:
        options = ChromeOptions()
        options.add_argument(f"--user-data-dir={profile_dir}")
        if headless:
            options.add_argument("--headless=new")
            options.add_argument("--disable-gpu")
            options.add_argument("--window-size=1920,1080")
            options.add_argument("--disable-dev-shm-usage")
        driver = webdriver.Chrome(options=options)

    if not headless:
        driver.maximize_window()
    else:
        try:
            driver.set_window_size(1920, 1080)
        except Exception:
            pass
    return driver


def login_manual_microsoft(
    driver: WebDriver, base_url: str, nombre_modalidad: str, headless: bool = False
) -> None:
    dashboard_url = f"{base_url}/my/"
    login_url = f"{base_url}/login/index.php"

    # Evita pedir ENTER cuando la sesion ya sigue activa en el perfil.
    try:
        driver.get(dashboard_url)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        current_url = driver.current_url or ""
        if sesion_activa_en_modalidad(current_url, base_url):
            print(f"[INFO] Sesion activa detectada en {nombre_modalidad}.")
            return
    except Exception:
        pass

    driver.get(login_url)
    time.sleep(1.0)

    # Algunos sitios redirigen automaticamente si ya hay cookie valida.
    current_url = driver.current_url or ""
    if sesion_activa_en_modalidad(current_url, base_url):
        print(f"[INFO] Sesion activa detectada en {nombre_modalidad} (redirigida).")
        return

    if headless:
        raise RuntimeError(
            f"No hay sesion activa en {nombre_modalidad} para modo headless. "
            "Ejecuta una corrida visible (sin headless) para autenticar."
        )

    print(
        f"\n[LOGIN {nombre_modalidad}] Inicia sesion en el navegador.\n"
        "1) Clic en 'Iniciar sesion con Microsoft'.\n"
        "2) Ingresa correo/contrasena.\n"
        "3) Completa autenticador.\n"
        "4) Si aparece, marca mantener sesion iniciada.\n"
        "5) Espera: el script detectara automaticamente cuando quedes dentro.\n"
    )

    # Espera activa del login para no pedir ENTER en cada corrida.
    deadline = time.time() + LOGIN_WAIT_SECONDS
    while time.time() < deadline:
        time.sleep(2.0)
        try:
            current_url = driver.current_url or ""
            if sesion_activa_en_modalidad(current_url, base_url):
                print(f"[INFO] Sesion iniciada en {nombre_modalidad}.")
                return
        except Exception:
            continue

    # Fallback: una ultima verificacion directa al dashboard.
    try:
        driver.get(dashboard_url)
        time.sleep(1.0)
        current_url = driver.current_url or ""
        if sesion_activa_en_modalidad(current_url, base_url):
            print(f"[INFO] Sesion iniciada en {nombre_modalidad} (verificacion final).")
            return
    except Exception:
        pass

    raise RuntimeError(
        f"No se detecto login completado en {nombre_modalidad} "
        f"despues de {LOGIN_WAIT_SECONDS} segundos."
    )


def elegir_mejor_link(course_links, query: str, titulos_fuente: Optional[Sequence[str]] = None):
    query_lower = query.lower()
    patt_exact = re.compile(rf"\b{re.escape(query_lower)}\b")
    titulos = [(t or "").strip() for t in (titulos_fuente or []) if (t or "").strip()]

    def score(link):
        text_raw = (link.text or "").strip()
        text = text_raw.lower()
        href = (link.get_attribute("href") or "").lower()
        if patt_exact.search(text):
            query_score = 3
        elif query_lower in text:
            query_score = 2
        elif query_lower in href:
            query_score = 1
        else:
            query_score = 0

        # En empates por NRC, prioriza el curso cuyo titulo coincide mejor con la fuente.
        title_score = 0.0
        if titulos:
            title_score = max(similarity_score(src, text_raw) for src in titulos)
        return (query_score, title_score)

    return max(course_links, key=score)


def buscar_curso_en_modalidad(
    driver: WebDriver,
    base_url: str,
    nrc_queries: Sequence[str],
    titulos_fuente: Optional[Sequence[str]] = None,
) -> Tuple[Optional[str], Optional[str], Optional[str], Optional[str], bool]:
    wait = WebDriverWait(driver, 12)
    for query in nrc_queries:
        search_url = f"{base_url}/course/search.php?areaids=core_course-course&q={query}"
        driver.get(search_url)
        time.sleep(1.5)

        current_url = driver.current_url or ""
        if not sesion_activa_en_modalidad(current_url, base_url) and url_indica_login(current_url):
            return None, None, None, None, True

        try:
            wait.until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
        except Exception:
            continue

        course_links = driver.find_elements(
            By.XPATH, "//a[contains(@href, '/course/view.php?id=')]"
        )
        if not course_links:
            continue

        best_link = elegir_mejor_link(course_links, query, titulos_fuente=titulos_fuente)
        href = best_link.get_attribute("href")
        nombre = (best_link.text or "").strip()
        course_id = obtener_id_desde_href(href)

        try:
            best_link.click()
            WebDriverWait(driver, 12).until(
                EC.url_contains("/course/view.php?id=")
            )
        except Exception:
            # Fallback: abrir por URL directa
            if href:
                driver.get(href)
                time.sleep(1.5)

        return course_id, nombre, href, query, False

    return None, None, None, None, False


def normalize_for_contains(text: str) -> str:
    text = text.lower()
    text = unicodedata.normalize("NFD", text)
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn")


def contar_mod_links(driver: WebDriver) -> int:
    script = """
        const links = [...document.querySelectorAll("a[href*='/mod/']")];
        const unique = new Set();
        for (const a of links) {
            const href = (a.getAttribute('href') || '').split('#')[0];
            if (href) unique.add(href);
        }
        return unique.size;
    """
    try:
        value = driver.execute_script(script)
        return int(value or 0)
    except Exception:
        return 0


def detectar_tipo_aula(page_source: str, mod_links: int) -> Tuple[str, str, str]:
    html = (page_source or "").lower()
    html_norm = normalize_for_contains(page_source or "")

    has_innovame = (
        'id="urlactivity"' in html
        or "uniminuto-uvd-format" in html
        or "course/format/uniminuto_uvd" in html
    )
    has_distancia40 = (
        'id="urlrecurso"' in html
        or "menu herramientas" in html_norm
        or re.search(r"format/uniminuto_course[s]?/", html) is not None
        or "uniminuto-formats" in html
        or "uniminuto-format single-section-format" in html
    )
    has_criba = (
        "format_onetopic-tabs" in html
        or ("course-content-header" in html and "-course-format" in html)
        or "onetopic-tab-body" in html
    )

    if mod_links <= VACIA_MAX_MOD_LINKS:
        return "vacia", f"mod_links={mod_links}<= {VACIA_MAX_MOD_LINKS}", "media"

    if has_innovame:
        return "innovame", "marcador_uniminuto_uvd/urlActivity", "alta"

    if has_distancia40:
        return "distancia 4.0", "marcador_urlRecurso/menu_herramientas", "alta"

    if has_criba:
        return "criba", "marcador_onetopic/course-format", "media"

    return "criba", "fallback_sin_marcadores_fuertes", "baja"


def detectar_no_matriculado(page_source: str) -> Optional[str]:
    html_norm = normalize_for_contains(page_source or "")
    for pattern in NO_MATRICULADO_PATTERNS:
        if pattern in html_norm:
            return pattern
    return None


def detectar_participantes_cero(
    driver: WebDriver,
    base_url: str,
    course_id: str,
) -> Tuple[Optional[int], bool, str]:
    if not base_url or not course_id:
        return None, False, ""

    target_url = f"{base_url}/user/index.php?id={course_id}"
    try:
        driver.get(target_url)
        time.sleep(1.2)
        current_url = driver.current_url or ""
        if url_indica_login(current_url):
            return None, False, "sin_acceso_participantes_o_relogin"

        body_text = ""
        try:
            body_text = driver.find_element(By.TAG_NAME, "body").text or ""
        except Exception:
            pass

        html_norm = normalize_for_contains(driver.page_source or "")
        body_norm = normalize_for_contains(body_text)
        for rgx in [r"\b(\d+)\s+participantes?\s+encontrados?\b", r"\b(\d+)\s+participants?\s+found\b"]:
            m = re.search(rgx, body_norm) or re.search(rgx, html_norm)
            if m:
                count = int(m.group(1))
                if count == 0:
                    return 0, True, f"{count} participantes encontrados"
                return count, False, f"{count} participantes encontrados"

        for pattern in PARTICIPANTES_CERO_PATTERNS:
            pattern_norm = normalize_for_contains(pattern)
            if pattern_norm in html_norm or pattern_norm in body_norm:
                return 0, True, pattern

        regex_variantes = [
            r"\b0\s+participantes?\s+encontrados?\b",
            r"\b0\s+participants?\s+found\b",
        ]
        for rgx in regex_variantes:
            if re.search(rgx, body_norm) or re.search(rgx, html_norm):
                return 0, True, "regex_0_participantes"
        return None, False, ""
    except Exception as e:
        return None, False, f"error_participantes: {e}"


def es_error_driver_desconectado(exc: Exception) -> bool:
    text = normalize_for_contains(str(exc))
    patrones = [
        "invalid session id",
        "session deleted",
        "no such window",
        "target window already closed",
        "web view not found",
        "not connected to devtools",
        "connection refused",
        "max retries exceeded",
        "failed to establish a new connection",
    ]
    return any(p in text for p in patrones)


def reiniciar_driver(
    browser: str, headless: bool = False, profile_dir_override: Optional[str] = None
) -> WebDriver:
    try:
        print(f"[INFO] Reabriendo {browser.upper()}...")
        return create_driver(
            browser=browser,
            headless=headless,
            profile_dir_override=profile_dir_override,
        )
    except Exception as e:
        print(f"[WARN] Fallo al reabrir {browser.upper()} con perfil actual: {e}")
        print(f"[INFO] Reintentando {browser.upper()} con perfil limpio...")
        return create_driver(
            browser=browser,
            clean_profile=True,
            headless=headless,
            profile_dir_override=profile_dir_override,
        )


def cargar_procesados(csv_path: Path) -> Set[str]:
    if not csv_path.exists():
        return set()
    done: Set[str] = set()
    with csv_path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh, delimiter=";")
        for row in reader:
            nrc = normalize_nrc(row.get("NRC", ""))
            if nrc:
                done.add(nrc)
    return done


def leer_headers_csv(csv_path: Path) -> List[str]:
    if not csv_path.exists():
        return []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh, delimiter=";")
        try:
            first = next(reader)
        except StopIteration:
            return []
    return [str(c or "").strip() for c in first]


def asegurar_esquema_checkpoint(csv_path: Path) -> None:
    if not csv_path.exists():
        return
    headers = leer_headers_csv(csv_path)
    expected = [str(h or "").strip() for h in OUTPUT_HEADERS]
    if headers == expected:
        return
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    backup = csv_path.with_name(f"{csv_path.stem}.schema_anterior_{timestamp}{csv_path.suffix}")
    shutil.move(str(csv_path), str(backup))
    print(
        f"[WARN] Checkpoint con esquema antiguo respaldado en: {backup}. "
        "Se creara uno nuevo con el formato actualizado."
    )


def sanitize_csv_value(value) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r\n", " | ").replace("\n", " | ").replace("\r", " | ")
    return text.strip()


def append_result_row(csv_path: Path, row: Dict[str, str]) -> None:
    headers = OUTPUT_HEADERS
    write_header = not csv_path.exists()
    clean_row = {k: sanitize_csv_value(row.get(k, "")) for k in headers}
    with csv_path.open("a", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers, delimiter=";")
        if write_header:
            writer.writeheader()
        writer.writerow(clean_row)


def append_csv_rows(source_csv: Path, target_csv: Path) -> int:
    if not source_csv.exists():
        return 0
    added = 0
    with source_csv.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh, delimiter=";")
        for row in reader:
            append_result_row(target_csv, row)
            added += 1
    return added


def generar_reporte_aulas_vacias(source_csv: Path, target_csv: Path) -> int:
    if target_csv.exists():
        target_csv.unlink()
    if not source_csv.exists():
        return 0

    count = 0
    with source_csv.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh, delimiter=";")
        for row in reader:
            es_vacia = str(row.get("ES_VACIA", "")).strip().upper() == "SI"
            es_vacia_participantes = str(row.get("ES_VACIA_POR_PARTICIPANTES", "")).strip().upper() == "SI"
            if not (es_vacia or es_vacia_participantes):
                continue
            append_result_row(target_csv, row)
            count += 1
    return count


def distribuir_en_workers(items: Sequence[str], workers: int) -> List[List[str]]:
    workers = max(1, workers)
    buckets: List[List[str]] = [[] for _ in range(workers)]
    for idx, item in enumerate(items):
        buckets[idx % workers].append(item)
    return buckets


def csv_to_xlsx(csv_path: Path, xlsx_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "TiposAula"

    with csv_path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh, delimiter=";")
        for row in reader:
            ws.append(row)

    if ws.max_row >= 1:
        header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            ws.column_dimensions[cell.column_letter].width = max(18, len(str(cell.value)) + 3)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    wb.save(xlsx_path)


def procesar_nrcs_lote(
    records: Dict[str, NRCRecord],
    nrcs: Sequence[str],
    browser: str,
    output_csv: Path,
    matricula_csv: Path,
    catalogo_d40: Sequence[str],
    headless: bool,
    worker_id: int = 1,
    total_workers: int = 1,
    profile_dir_override: Optional[str] = None,
    strict_modalidad: bool = False,
    modalidades_forzadas: Optional[Sequence[str]] = None,
    prelogin_modalidades: Optional[Sequence[str]] = None,
    solo_nrc_5_digitos: bool = False,
    nrc_5_segun_archivo: bool = False,
) -> int:
    nrcs_list = list(nrcs)
    if not nrcs_list:
        print(f"[W{worker_id}/{total_workers}] Sin NRC asignados.")
        return 0

    pref = f"[W{worker_id}/{total_workers}]"
    try:
        driver = create_driver(
            browser=browser,
            headless=headless,
            profile_dir_override=profile_dir_override,
        )
    except RuntimeError as e:
        print(f"{pref} [ERROR] {e}")
        return 0
    except Exception as e:
        print(f"{pref} [WARN] Fallo al abrir {browser.upper()} con perfil actual: {e}")
        print(f"{pref} [INFO] Reintentando con perfil limpio...")
        try:
            driver = create_driver(
                browser=browser,
                clean_profile=True,
                headless=headless,
                profile_dir_override=profile_dir_override,
            )
        except Exception as e2:
            print(f"{pref} [ERROR] No se pudo abrir {browser.upper()}: {e2}")
            return 0

    modalidades_forzadas = unique_keep_order(modalidades_forzadas or [])
    prelogin_modalidades = unique_keep_order(prelogin_modalidades or [])
    sesiones_activas: Set[str] = set()
    filas_escritas = 0
    try:
        if prelogin_modalidades:
            print(
                f"{pref} [INFO] Pre-login inicial: "
                + ", ".join(MODALIDADES[m]["nombre"] for m in prelogin_modalidades if m in MODALIDADES)
            )
            for mod in prelogin_modalidades:
                cfg = MODALIDADES.get(mod)
                if not cfg:
                    continue
                login_manual_microsoft(
                    driver,
                    cfg["base_url"],
                    cfg["nombre"],
                    headless=headless,
                )
                sesiones_activas.add(mod)

        for idx, nrc in enumerate(nrcs_list, start=1):
            rec = records[nrc]
            periodos = ", ".join(sorted(p for p in rec.periodos if p))
            metodos = ", ".join(sorted(rec.metodos))
            archivos = " | ".join(sorted(rec.archivos))

            print("\n" + "=" * 80)
            print(
                f"{pref} [{idx}/{len(nrcs_list)}] NRC {nrc} | "
                f"Periodos: {periodos} | Metodos: {metodos}"
            )
            nrc_procesado = False

            for intento in range(1, MAX_REINTENTOS_NRC + 1):
                found = False
                course_id = None
                nombre_curso = None
                consulta_usada = None
                modalidad_encontrada = None
                base_url_encontrada = ""
                tipo_aula = ""
                confianza = ""
                regla = ""
                mod_links = 0
                error_msg = ""
                usar_solo_nrc_5 = solo_nrc_5_digitos or (
                    nrc_5_segun_archivo and requiere_solo_nrc_5_por_archivo(rec)
                )
                nrc_queries = construir_nrc_queries(rec, nrc, usar_solo_nrc_5)

                if modalidades_forzadas:
                    modalidades_busqueda = list(modalidades_forzadas)
                elif strict_modalidad:
                    modalidades_busqueda = unique_keep_order(rec.modalidades_preferidas)
                    if not modalidades_busqueda:
                        modalidades_busqueda = list(TODAS_MODALIDADES)
                else:
                    modalidades_busqueda = unique_keep_order(rec.modalidades_preferidas + TODAS_MODALIDADES)

                try:
                    for mod in modalidades_busqueda:
                        cfg = MODALIDADES[mod]
                        if mod not in sesiones_activas:
                            login_manual_microsoft(
                                driver,
                                cfg["base_url"],
                                cfg["nombre"],
                                headless=headless,
                            )
                            sesiones_activas.add(mod)

                        course_id, nombre_curso, _href, consulta_usada, need_login = buscar_curso_en_modalidad(
                            driver,
                            cfg["base_url"],
                            nrc_queries,
                            titulos_fuente=sorted(rec.titulos),
                        )

                        if need_login:
                            print(
                                f"{pref} [WARN] Sesion expirada en {cfg['nombre']}. "
                                "Reautenticando..."
                            )
                            login_manual_microsoft(
                                driver,
                                cfg["base_url"],
                                cfg["nombre"],
                                headless=headless,
                            )
                            sesiones_activas.add(mod)
                            course_id, nombre_curso, _href, consulta_usada, _ = buscar_curso_en_modalidad(
                                driver,
                                cfg["base_url"],
                                nrc_queries,
                                titulos_fuente=sorted(rec.titulos),
                            )

                        if course_id:
                            modalidad_encontrada = cfg["nombre"]
                            base_url_encontrada = cfg["base_url"]
                            found = True
                            break

                    if not found:
                        print(f"{pref} [WARN] NRC {nrc} no encontrado en ninguna modalidad.")
                        estado_no_encontrado = (
                            "NO_ENCONTRADO_MODALIDAD_OBJETIVO"
                            if (strict_modalidad or modalidades_forzadas)
                            else "NO_ENCONTRADO"
                        )
                        error_no_encontrado = ""
                        if strict_modalidad or modalidades_forzadas:
                            modalidades_txt = ", ".join(
                                MODALIDADES[m]["nombre"]
                                for m in modalidades_busqueda
                                if m in MODALIDADES
                            )
                            if modalidades_txt:
                                error_no_encontrado = f"modalidades_intentadas={modalidades_txt}"
                        append_result_row(
                            output_csv,
                            build_enriched_row(
                                rec=rec,
                                nrc=nrc,
                                tipo_aula="",
                                total_participantes=None,
                                es_vacia_por_participantes=False,
                                texto_participantes="",
                                confianza="",
                                regla="",
                                mod_links="",
                                modalidad_encontrada="",
                                course_id="",
                                nombre_curso_moodle="",
                                consulta_usada="",
                                periodos=periodos,
                                metodos=metodos,
                                archivos=archivos,
                                estado=estado_no_encontrado,
                                error=error_no_encontrado,
                                catalogo_d40=catalogo_d40,
                            ),
                        )
                        filas_escritas += 1
                        nrc_procesado = True
                        break

                    time.sleep(1.0)
                    no_matriculado_pattern = detectar_no_matriculado(driver.page_source)
                    if no_matriculado_pattern:
                        error_msg = f"no_matriculado: '{no_matriculado_pattern}'"
                        print(
                            f"{pref} [WARN] NRC {nrc} | Modalidad: {modalidad_encontrada} | "
                            "Curso encontrado pero sin matriculacion."
                        )
                        row = build_enriched_row(
                            rec=rec,
                            nrc=nrc,
                            tipo_aula="",
                            total_participantes=None,
                            es_vacia_por_participantes=False,
                            texto_participantes="",
                            confianza="alta",
                            regla="texto_no_matriculado",
                            mod_links="",
                            modalidad_encontrada=modalidad_encontrada or "",
                            course_id=course_id or "",
                            nombre_curso_moodle=nombre_curso or "",
                            consulta_usada=consulta_usada or "",
                            periodos=periodos,
                            metodos=metodos,
                            archivos=archivos,
                            estado="SIN_MATRICULA",
                            error=error_msg,
                            catalogo_d40=catalogo_d40,
                        )
                        append_result_row(output_csv, row)
                        append_result_row(matricula_csv, row)
                        filas_escritas += 1
                        nrc_procesado = True
                        break

                    mod_links = contar_mod_links(driver)
                    tipo_aula, regla, confianza = detectar_tipo_aula(driver.page_source, mod_links)
                    total_participantes = None
                    es_vacia_por_participantes = False
                    texto_participantes = ""
                    if course_id and base_url_encontrada:
                        total_participantes, es_vacia_por_participantes, texto_participantes = detectar_participantes_cero(
                            driver=driver,
                            base_url=base_url_encontrada,
                            course_id=course_id,
                        )
                        print(
                            f"{pref} [INFO] NRC {nrc} | Participantes: "
                            + (str(total_participantes) if total_participantes is not None else "N/D")
                            + " | Vacia por participantes: "
                            + ("SI" if es_vacia_por_participantes else "NO")
                            + (f" ({texto_participantes})" if texto_participantes else "")
                        )
                    print(
                        f"{pref} [OK] NRC {nrc} | Modalidad: {modalidad_encontrada} | "
                        f"Tipo: {tipo_aula} | mod_links={mod_links}"
                    )

                    append_result_row(
                        output_csv,
                        build_enriched_row(
                            rec=rec,
                            nrc=nrc,
                            tipo_aula=tipo_aula,
                            total_participantes=total_participantes,
                            es_vacia_por_participantes=es_vacia_por_participantes,
                            texto_participantes=texto_participantes,
                            confianza=confianza,
                            regla=regla,
                            mod_links=str(mod_links),
                            modalidad_encontrada=modalidad_encontrada or "",
                            course_id=course_id or "",
                            nombre_curso_moodle=nombre_curso or "",
                            consulta_usada=consulta_usada or "",
                            periodos=periodos,
                            metodos=metodos,
                            archivos=archivos,
                            estado="OK",
                            error="",
                            catalogo_d40=catalogo_d40,
                        ),
                    )
                    filas_escritas += 1
                    nrc_procesado = True
                    break

                except Exception as e:
                    if es_error_driver_desconectado(e) and intento < MAX_REINTENTOS_NRC:
                        print(
                            f"{pref} [WARN] Driver desconectado procesando NRC {nrc}. "
                            f"Reintentando ({intento}/{MAX_REINTENTOS_NRC})..."
                        )
                        try:
                            driver.quit()
                        except Exception:
                            pass
                        driver = reiniciar_driver(
                            browser,
                            headless=headless,
                            profile_dir_override=profile_dir_override,
                        )
                        sesiones_activas.clear()
                        time.sleep(1.0)
                        continue

                    error_msg = str(e)
                    print(f"{pref} [WARN] Error procesando NRC {nrc}: {error_msg}")
                    estado_error = "ERROR_EJECUCION"
                    if headless and "No hay sesion activa en" in error_msg:
                        estado_error = "ERROR_LOGIN_REQUERIDO"
                    append_result_row(
                        output_csv,
                        build_enriched_row(
                            rec=rec,
                            nrc=nrc,
                            tipo_aula="",
                            total_participantes=None,
                            es_vacia_por_participantes=False,
                            texto_participantes="",
                            confianza="",
                            regla="",
                            mod_links="",
                            modalidad_encontrada=modalidad_encontrada or "",
                            course_id=course_id or "",
                            nombre_curso_moodle=nombre_curso or "",
                            consulta_usada=consulta_usada or "",
                            periodos=periodos,
                            metodos=metodos,
                            archivos=archivos,
                            estado=estado_error,
                            error=error_msg,
                            catalogo_d40=catalogo_d40,
                        ),
                    )
                    filas_escritas += 1
                    nrc_procesado = True
                    if estado_error == "ERROR_LOGIN_REQUERIDO":
                        for nrc_restante in nrcs_list[idx:]:
                            rec_rest = records[nrc_restante]
                            append_result_row(
                                output_csv,
                                build_enriched_row(
                                    rec=rec_rest,
                                    nrc=nrc_restante,
                                    tipo_aula="",
                                    total_participantes=None,
                                    es_vacia_por_participantes=False,
                                    texto_participantes="",
                                    confianza="",
                                    regla="",
                                    mod_links="",
                                    modalidad_encontrada="",
                                    course_id="",
                                    nombre_curso_moodle="",
                                    consulta_usada="",
                                    periodos=", ".join(sorted(p for p in rec_rest.periodos if p)),
                                    metodos=", ".join(sorted(rec_rest.metodos)),
                                    archivos=" | ".join(sorted(rec_rest.archivos)),
                                    estado="ERROR_LOGIN_REQUERIDO",
                                    error=error_msg,
                                    catalogo_d40=catalogo_d40,
                                ),
                            )
                            filas_escritas += 1
                        print(
                            f"{pref} [ERROR] Worker detenido por falta de sesion en modo headless. "
                            "Ejecuta una corrida visible para autenticar."
                        )
                        return filas_escritas
                    break

            if not nrc_procesado:
                append_result_row(
                    output_csv,
                    build_enriched_row(
                        rec=rec,
                        nrc=nrc,
                        tipo_aula="",
                        total_participantes=None,
                        es_vacia_por_participantes=False,
                        texto_participantes="",
                        confianza="",
                        regla="",
                        mod_links="",
                        modalidad_encontrada="",
                        course_id="",
                        nombre_curso_moodle="",
                        consulta_usada="",
                        periodos=periodos,
                        metodos=metodos,
                        archivos=archivos,
                        estado="ERROR_SIN_DETALLE",
                        error="",
                        catalogo_d40=catalogo_d40,
                    ),
                )
                filas_escritas += 1
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    return filas_escritas


def run(args):
    input_dir = Path(args.input_dir)
    output_xlsx = Path(args.output)
    output_csv = output_xlsx.with_suffix(".csv")
    matricula_csv = output_xlsx.with_name(f"{output_xlsx.stem}_SIN_MATRICULA.csv")
    matricula_xlsx = output_xlsx.with_name(f"{output_xlsx.stem}_SIN_MATRICULA.xlsx")
    vacias_csv = output_xlsx.with_name(f"{output_xlsx.stem}_AULAS_VACIAS.csv")
    vacias_xlsx = output_xlsx.with_name(f"{output_xlsx.stem}_AULAS_VACIAS.xlsx")
    workers = max(1, min(int(args.workers or 1), MAX_WORKERS))
    if workers != int(args.workers or 1):
        print(f"[WARN] workers ajustado a {workers} (rango permitido: 1-{MAX_WORKERS}).")
    prioridad_periodos = [
        normalize_periodo(p)
        for p in (args.prioridad_periodos or "").split(",")
        if normalize_periodo(p)
    ]
    modalidades_permitidas = parse_modalidades_cli(args.modalidades_permitidas)
    prelogin_modalidades = parse_modalidades_cli(args.prelogin_modalidades)
    if args.prelogin_all_modalidades:
        prelogin_modalidades = list(TODAS_MODALIDADES)
    if modalidades_permitidas:
        prelogin_modalidades = unique_keep_order(prelogin_modalidades + modalidades_permitidas)
    catalogo_path = Path(args.catalogo_distancia_xlsx)
    catalogo_col = max(1, int(args.catalogo_distancia_columna or DEFAULT_CATALOGO_D40_COL))
    if catalogo_col != int(args.catalogo_distancia_columna or DEFAULT_CATALOGO_D40_COL):
        print(f"[WARN] Columna de catalogo ajustada a {catalogo_col}.")
    catalogo_d40 = cargar_catalogo_distancia40(
        xlsx_path=catalogo_path,
        sheet_name=args.catalogo_distancia_hoja,
        col_index=catalogo_col,
    )
    print(f"[INFO] Catalogo Distancia 4.0: {catalogo_path}")
    print(f"[INFO] Cursos en catalogo Distancia 4.0: {len(catalogo_d40)}")

    records = cargar_nrcs(input_dir)
    if not records:
        raise RuntimeError("No hay NRC para procesar.")

    asegurar_esquema_checkpoint(output_csv)
    asegurar_esquema_checkpoint(matricula_csv)
    processed = cargar_procesados(output_csv) if args.resume else set()

    def sort_key(nrc: str) -> Tuple[int, int]:
        rec = records[nrc]
        rec_periodos = {normalize_periodo(p) for p in rec.periodos if p}
        rank = len(prioridad_periodos)
        for i, periodo_prioritario in enumerate(prioridad_periodos):
            if periodo_prioritario in rec_periodos:
                rank = i
                break
        return (rank, int(nrc))

    pendientes = [nrc for nrc in sorted(records.keys(), key=sort_key) if nrc not in processed]

    print(f"[INFO] NRC totales: {len(records)}")
    print(f"[INFO] NRC ya procesados: {len(processed)}")
    print(f"[INFO] NRC pendientes: {len(pendientes)}")
    if prioridad_periodos:
        print(f"[INFO] Prioridad de periodos: {', '.join(prioridad_periodos)}")
    print(f"[INFO] CSV checkpoint: {output_csv}")
    print(f"[INFO] Excel final: {output_xlsx}")
    print(f"[INFO] Reporte sin matricula: {matricula_xlsx}")
    print(f"[INFO] Reporte aulas vacias/sin participantes: {vacias_xlsx}")
    print(f"[INFO] Headless: {'SI' if args.headless else 'NO'}")
    print(f"[INFO] Workers: {workers}")
    print(
        f"[INFO] Modo estricto por modalidad: "
        f"{'SI' if args.modo_estricto_modalidad else 'NO'}"
    )
    print(
        f"[INFO] Consulta NRC solo 5 digitos: "
        f"{'SI' if args.solo_nrc_5_digitos else 'NO'}"
    )
    print(
        f"[INFO] Consulta NRC 5 digitos segun archivo/modalidad: "
        f"{'SI' if args.nrc_5_segun_archivo else 'NO'}"
    )
    if modalidades_permitidas:
        print(
            "[INFO] Modalidades permitidas: "
            + ", ".join(MODALIDADES[m]["nombre"] for m in modalidades_permitidas if m in MODALIDADES)
        )
    if prelogin_modalidades:
        print(
            "[INFO] Pre-login inicial de modalidades: "
            + ", ".join(MODALIDADES[m]["nombre"] for m in prelogin_modalidades if m in MODALIDADES)
        )

    if not pendientes:
        print("[INFO] No hay pendientes. Regenerando Excel...")
        if output_csv.exists():
            csv_to_xlsx(output_csv, output_xlsx)
        if matricula_csv.exists():
            csv_to_xlsx(matricula_csv, matricula_xlsx)
        total_vacias = generar_reporte_aulas_vacias(output_csv, vacias_csv)
        if total_vacias > 0:
            csv_to_xlsx(vacias_csv, vacias_xlsx)
        else:
            if vacias_csv.exists():
                vacias_csv.unlink()
            if vacias_xlsx.exists():
                vacias_xlsx.unlink()
        print("[OK] Excel actualizado.")
        return

    if workers == 1:
        perfiles = {1: str(profile_dir_base(args.browser))}
        procesados_now = procesar_nrcs_lote(
            records=records,
            nrcs=pendientes,
            browser=args.browser,
            output_csv=output_csv,
            matricula_csv=matricula_csv,
            catalogo_d40=catalogo_d40,
            headless=args.headless,
            worker_id=1,
            total_workers=1,
            profile_dir_override=perfiles[1],
            strict_modalidad=args.modo_estricto_modalidad,
            modalidades_forzadas=modalidades_permitidas,
            prelogin_modalidades=prelogin_modalidades,
            solo_nrc_5_digitos=args.solo_nrc_5_digitos,
            nrc_5_segun_archivo=args.nrc_5_segun_archivo,
        )
        print(f"[INFO] Registros procesados en esta corrida: {procesados_now}")
    else:
        print("[INFO] Modo multi-worker activado.")
        if not args.headless:
            print(
                "[WARN] Ejecutar varios workers sin headless abrira multiples ventanas del navegador."
            )

        perfiles = preparar_perfiles_workers(args.browser, workers)
        buckets = distribuir_en_workers(pendientes, workers)
        temp_dir = output_xlsx.parent / f".{output_xlsx.stem}_workers"
        temp_dir.mkdir(parents=True, exist_ok=True)

        jobs: List[Tuple[int, List[str], Path, Path]] = []
        for worker_id, lote in enumerate(buckets, start=1):
            part_csv = temp_dir / f"worker_{worker_id}.csv"
            part_matricula = temp_dir / f"worker_{worker_id}_SIN_MATRICULA.csv"
            if part_csv.exists():
                part_csv.unlink()
            if part_matricula.exists():
                part_matricula.unlink()
            jobs.append((worker_id, lote, part_csv, part_matricula))

        with ThreadPoolExecutor(max_workers=workers) as executor:
            future_map = {
                executor.submit(
                    procesar_nrcs_lote,
                    records,
                    lote,
                    args.browser,
                    part_csv,
                    part_matricula,
                    catalogo_d40,
                    args.headless,
                    worker_id,
                    workers,
                    perfiles.get(worker_id),
                    args.modo_estricto_modalidad,
                    modalidades_permitidas,
                    prelogin_modalidades,
                    args.solo_nrc_5_digitos,
                    args.nrc_5_segun_archivo,
                ): (worker_id, len(lote))
                for worker_id, lote, part_csv, part_matricula in jobs
            }
            for future in as_completed(future_map):
                worker_id, total_lote = future_map[future]
                try:
                    procesados_worker = future.result()
                    print(
                        f"[INFO] Worker {worker_id} finalizado: "
                        f"{procesados_worker}/{total_lote} registros."
                    )
                except Exception as e:
                    print(f"[ERROR] Worker {worker_id} fallo: {e}")

        total_merge = 0
        total_merge_matricula = 0
        for _, _, part_csv, part_matricula in jobs:
            total_merge += append_csv_rows(part_csv, output_csv)
            total_merge_matricula += append_csv_rows(part_matricula, matricula_csv)
        print(f"[INFO] Filas consolidadas al checkpoint: {total_merge}")
        print(f"[INFO] Filas consolidadas en sin matricula: {total_merge_matricula}")

    if output_csv.exists():
        csv_to_xlsx(output_csv, output_xlsx)
    if matricula_csv.exists():
        csv_to_xlsx(matricula_csv, matricula_xlsx)
    total_vacias = generar_reporte_aulas_vacias(output_csv, vacias_csv)
    if total_vacias > 0:
        csv_to_xlsx(vacias_csv, vacias_xlsx)
    else:
        if vacias_csv.exists():
            vacias_csv.unlink()
        if vacias_xlsx.exists():
            vacias_xlsx.unlink()
    print("\n[FIN] Proceso completado.")
    print(f"[OK] CSV: {output_csv}")
    print(f"[OK] Excel: {output_xlsx}")
    if matricula_csv.exists():
        print(f"[OK] CSV sin matricula: {matricula_csv}")
        print(f"[OK] Excel sin matricula: {matricula_xlsx}")
    if vacias_csv.exists():
        print(f"[OK] CSV aulas vacias/sin participantes: {vacias_csv} ({total_vacias} filas)")
        print(f"[OK] Excel aulas vacias/sin participantes: {vacias_xlsx}")
    else:
        print("[INFO] No se encontraron aulas vacias/sin participantes para reporte dedicado.")
    if args.pause_at_end and not args.headless:
        input("\nPresiona ENTER para cerrar el navegador...\n")


def build_args():
    parser = argparse.ArgumentParser(
        description="Clasifica tipos de aula por NRC directamente en Moodle y genera Excel."
    )
    parser.add_argument(
        "--browser",
        default="edge",
        choices=["edge", "chrome"],
        help="Navegador para automatizacion Selenium (default: edge).",
    )
    parser.add_argument(
        "--input-dir",
        default=str(DEFAULT_INPUT_DIR),
        help="Directorio con los CSV de programacion academica.",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT_XLSX),
        help="Ruta del Excel de salida.",
    )
    parser.add_argument(
        "--prioridad-periodos",
        default="202615",
        help="Lista separada por comas de periodos a priorizar al inicio (default: 202615).",
    )
    parser.add_argument(
        "--catalogo-distancia-xlsx",
        default=str(DEFAULT_CATALOGO_D40_XLSX),
        help="Ruta del catalogo Excel con cursos Distancia 4.0.",
    )
    parser.add_argument(
        "--catalogo-distancia-hoja",
        default=DEFAULT_CATALOGO_D40_HOJA,
        help="Hoja del catalogo Distancia 4.0 (default: DISTANCIA 4.0).",
    )
    parser.add_argument(
        "--catalogo-distancia-columna",
        type=int,
        default=DEFAULT_CATALOGO_D40_COL,
        help="Indice 1-based de columna con nombres en el catalogo (default: 3, columna C).",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Ejecuta el navegador oculto (sin ventana visible).",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=1,
        help="Cantidad de workers en paralelo (default: 1, maximo interno: 8).",
    )
    parser.add_argument(
        "--modo-estricto-modalidad",
        action="store_true",
        help=(
            "Busca cada NRC solo en la(s) modalidad(es) inferida(s) por archivo/periodo "
            "(sin fallback automatico a otras modalidades)."
        ),
    )
    parser.add_argument(
        "--modalidades-permitidas",
        default="",
        help=(
            "Restringe la busqueda a modalidades especificas (coma separadas). "
            "Ej: distancia,presencial"
        ),
    )
    parser.add_argument(
        "--solo-nrc-5-digitos",
        action="store_true",
        help="Consulta solo el NRC normalizado a 5 digitos (sin quitar ceros a la izquierda).",
    )
    parser.add_argument(
        "--nrc-5-segun-archivo",
        action="store_true",
        help=(
            "Usa solo NRC de 5 digitos para archivos de pregrado/posgrado "
            "presencial o distancia; en los demas permite tambien NRC sin ceros."
        ),
    )
    parser.add_argument(
        "--prelogin-all-modalidades",
        action="store_true",
        help="Al iniciar cada worker, solicita login en todas las modalidades antes de procesar NRC.",
    )
    parser.add_argument(
        "--prelogin-modalidades",
        default="",
        help=(
            "Lista coma separada de modalidades para pre-login inicial. "
            "Ej: distancia,moocs"
        ),
    )
    parser.add_argument(
        "--resume",
        action="store_true",
        default=True,
        help="Reanuda desde CSV checkpoint si ya existe (default: activado).",
    )
    parser.add_argument(
        "--no-resume",
        action="store_true",
        help="Ignora checkpoint y reprocesa todos los NRC.",
    )
    parser.add_argument(
        "--pause-at-end",
        action="store_true",
        help="Pausa al final para cerrar el navegador manualmente.",
    )
    return parser.parse_args()


def main():
    args = build_args()
    if args.no_resume:
        args.resume = False
    run(args)


if __name__ == "__main__":
    main()
